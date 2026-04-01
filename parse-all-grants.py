#!/usr/bin/env python3
"""
Parse all downloaded grants data from multiple sources into a single CSV.

Sources:
- 360Giving format CSVs (Camden, Trafford)
- 360Giving format XLSX (Birmingham, Barnet, Essex)
- Council XLSX (South Oxfordshire, Vale of White Horse)
- Spending CSVs with grant extraction (Cambridgeshire, Epping Forest)

Output: parsed-grants.csv
"""

import csv
import os
from collections import defaultdict
from datetime import datetime

import openpyxl

DIR = "src/data/councils/pdfs/spending-csvs"


def parse_360giving_csv(filepath, council_name, min_year=2020):
    """Parse a 360Giving standard CSV file."""
    grants = defaultdict(lambda: {"total": 0, "count": 0, "purpose": ""})

    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    recipient = (row.get("Recipient Org:Name", "") or "").strip()
                    amt_str = (row.get("Amount Awarded", "") or row.get("Amount Disbursed", "") or "").strip()
                    date_str = row.get("Award Date", "") or ""
                    desc = (row.get("Description", "") or row.get("Title", "") or "").strip()

                    if not recipient or not amt_str:
                        continue
                    try:
                        amt = abs(float(amt_str))
                    except (ValueError, TypeError):
                        continue

                    if amt < 100:
                        continue

                    # Filter by year
                    try:
                        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
                        if dt.year < min_year:
                            continue
                    except (ValueError, TypeError):
                        pass  # Keep if date unparseable

                    grants[recipient]["total"] += amt
                    grants[recipient]["count"] += 1
                    if not grants[recipient]["purpose"] and desc:
                        grants[recipient]["purpose"] = desc[:200]

            return council_name, dict(grants)
        except (UnicodeDecodeError, csv.Error):
            continue

    return council_name, {}


def parse_360giving_xlsx(filepath, council_name, min_year=2020):
    """Parse a 360Giving standard XLSX file."""
    grants = defaultdict(lambda: {"total": 0, "count": 0, "purpose": ""})

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h_map = {h.lower(): i for i, h in enumerate(headers)}

    recipient_col = None
    amount_col = None
    date_col = None
    desc_col = None

    for h, i in h_map.items():
        if "recipient" in h and "name" in h and recipient_col is None:
            recipient_col = i
        elif "amount awarded" in h and amount_col is None:
            amount_col = i
        elif "amount disbursed" in h and amount_col is None:
            amount_col = i
        elif "award date" in h and date_col is None:
            date_col = i
        elif "description" in h and desc_col is None:
            desc_col = i
        elif "title" in h and desc_col is None:
            desc_col = i

    if recipient_col is None or amount_col is None:
        # Try council-specific formats (South Oxfordshire style)
        for h, i in h_map.items():
            if "organisation awarded" in h:
                recipient_col = i
            elif "grant value" in h:
                amount_col = i
            elif "offer" in h and "date" in h:
                date_col = i
            elif "description" in h:
                desc_col = i

    if recipient_col is None or amount_col is None:
        wb.close()
        return council_name, {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(recipient_col, amount_col):
            continue

        recipient = str(row[recipient_col] or "").strip()
        amt = row[amount_col]
        date_val = row[date_col] if date_col is not None and date_col < len(row) else None
        desc = str(row[desc_col] or "").strip() if desc_col is not None and desc_col < len(row) else ""

        if not recipient or not amt:
            continue
        try:
            amt = abs(float(amt))
        except (ValueError, TypeError):
            continue
        if amt < 100:
            continue

        # Filter by year
        date_str = str(date_val or "")
        try:
            if isinstance(date_val, datetime):
                if date_val.year < min_year:
                    continue
            elif len(date_str) >= 4:
                year = int(date_str[:4])
                if year < min_year:
                    continue
        except (ValueError, TypeError):
            pass

        grants[recipient]["total"] += amt
        grants[recipient]["count"] += 1
        if not grants[recipient]["purpose"] and desc:
            grants[recipient]["purpose"] = desc[:200]

    wb.close()
    return council_name, dict(grants)


def parse_spending_csv_for_grants(filepath, council_name):
    """Parse a spending-over-£500 CSV and extract grant payments."""
    grants = defaultdict(lambda: {"total": 0, "count": 0, "purpose": ""})

    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                first = f.readline()
                if "<html" in first.lower():
                    return council_name, {}
                f.seek(0)

                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                h_lower = {h.lower().strip(): h for h in headers}

                # Find columns
                amount_col = supplier_col = expense_col = supplier_type_col = None
                for key, orig in h_lower.items():
                    if "amount" in key and not amount_col:
                        amount_col = orig
                    elif ("supplier" in key or "vendor" in key or "payee" in key) and "type" not in key and not supplier_col:
                        supplier_col = orig
                    elif ("expense" in key or "category" in key) and "type" in key and not expense_col:
                        expense_col = orig
                    elif "supplier type" in key:
                        supplier_type_col = orig

                if not supplier_col or not amount_col:
                    continue

                for row in reader:
                    supplier = (row.get(supplier_col, "") or "").strip()
                    if not supplier or "redacted" in supplier.lower() or "personal" in supplier.lower():
                        continue

                    expense = (row.get(expense_col, "") or "").lower() if expense_col else ""
                    supplier_type = (row.get(supplier_type_col, "") or "").lower() if supplier_type_col else ""

                    # Only genuine voluntary sector grants
                    is_grant = False
                    if expense in ["grants", "expenditure on grants", "grants to voluntary orgs",
                                   "grants to voluntary organisations", "voluntary sector grants",
                                   "grants & contributions", "contributions to partnerships"]:
                        is_grant = True
                    elif supplier_type in ["charity", "local clubs & groups", "trust (other)", "voluntary"]:
                        if "grant" in expense:
                            is_grant = True

                    if not is_grant:
                        continue

                    # Skip commercial entities
                    sl = supplier.lower()
                    if any(x in sl for x in ["nhs", " plc", "redacted", "hmrc", "police", "fire"]):
                        continue

                    amt_str = (row.get(amount_col, "") or "").replace(",", "").replace("£", "").replace("\xa3", "").replace('"', "").strip()
                    try:
                        amt = abs(float(amt_str))
                    except (ValueError, TypeError):
                        continue
                    if amt < 100 or amt > 10_000_000:
                        continue

                    grants[supplier]["total"] += amt
                    grants[supplier]["count"] += 1
                    if not grants[supplier]["purpose"]:
                        grants[supplier]["purpose"] = expense.title()

            return council_name, dict(grants)
        except (UnicodeDecodeError, csv.Error):
            continue

    return council_name, {}


def main():
    print("=== Parse All Grants Data ===\n")

    all_results = {}

    # 360Giving CSVs
    for council, filename in [
        ("Camden", "camden-grants.csv"),
        ("Trafford", "trafford-grants.csv"),
    ]:
        path = f"{DIR}/{filename}"
        if os.path.exists(path):
            name, grants = parse_360giving_csv(path, council)
            if grants:
                all_results[name] = grants
                print(f"  {name}: {len(grants)} recipients (360Giving CSV)")

    # 360Giving XLSX
    for council, filename in [
        ("Birmingham", "birmingham-grants.xlsx"),
        ("Barnet", "barnet-grants.xlsx"),
        ("Essex", "essex-grants.xlsx"),
    ]:
        path = f"{DIR}/{filename}"
        if os.path.exists(path):
            name, grants = parse_360giving_xlsx(path, council)
            if grants:
                all_results[name] = grants
                print(f"  {name}: {len(grants)} recipients (360Giving XLSX)")

    # Council XLSX (South Oxfordshire format)
    for council, filename in [
        ("South Oxfordshire", "south-oxfordshire-grants.xlsx"),
        ("Vale of White Horse", "vale-of-white-horse-grants.xlsx"),
    ]:
        path = f"{DIR}/{filename}"
        if os.path.exists(path):
            name, grants = parse_360giving_xlsx(path, council, min_year=2022)
            if grants:
                all_results[name] = grants
                print(f"  {name}: {len(grants)} recipients (council XLSX)")

    # Spending CSVs with grant extraction
    for council, filename in [
        ("Cambridgeshire", "cambridgeshire.csv"),
    ]:
        path = f"{DIR}/{filename}"
        if os.path.exists(path):
            name, grants = parse_spending_csv_for_grants(path, council)
            if grants:
                all_results[name] = grants
                print(f"  {name}: {len(grants)} recipients (spending CSV)")

    # Epping Forest (already downloaded from earlier run)
    ef_path = f"{DIR}/epping-forest.csv"
    if os.path.exists(ef_path):
        name, grants = parse_spending_csv_for_grants(ef_path, "Epping Forest")
        if grants:
            all_results[name] = grants
            print(f"  Epping Forest: {len(grants)} recipients (spending CSV)")

    print(f"\n=== Results ===")
    print(f"Councils with grants data: {len(all_results)}")

    # Write output CSV
    with open("parsed-grants.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["council", "rank", "recipient", "amount", "purpose"])

        for council in sorted(all_results.keys()):
            grants = all_results[council]
            sorted_grants = sorted(grants.items(), key=lambda x: x[1]["total"], reverse=True)[:15]
            for rank, (recipient, data) in enumerate(sorted_grants, 1):
                writer.writerow([
                    council, rank, recipient,
                    round(data["total"]),
                    data.get("purpose", "")[:200],
                ])

    total_rows = sum(min(len(g), 15) for g in all_results.values())
    print(f"Saved {total_rows} rows to parsed-grants.csv")

    # Show examples
    for council in sorted(all_results.keys()):
        print(f"\n{council} - Top 5 grants:")
        grants = all_results[council]
        sorted_g = sorted(grants.items(), key=lambda x: x[1]["total"], reverse=True)[:5]
        for name, data in sorted_g:
            print(f"  {name[:45]:45s} £{data['total']:>10,.0f}")


if __name__ == "__main__":
    main()
